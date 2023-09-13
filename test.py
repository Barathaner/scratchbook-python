import cv2
import numpy as np
import random
from openpyxl import Workbook

def generate_random_color():
    return (random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))

def draw_lines_between_gaps(image, contours, max_gap=20):
    for contour in contours:
        for i in range(len(contour)):
            point1 = tuple(contour[i][0])
            point2 = tuple(contour[(i+1)%len(contour)][0])
            if cv2.norm(np.array(point1)-np.array(point2)) > max_gap:
                cv2.line(image, point1, point2, (0, 0, 0), 2)

# Bild laden
image = cv2.imread('grundriss.jpg', 0)
edges = cv2.Canny(image, 50, 150)
contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
draw_lines_between_gaps(edges, contours)

kernel = np.ones((5, 5), np.uint8)
dilated_image = cv2.dilate(edges, kernel, iterations=3)
contours, _ = cv2.findContours(dilated_image, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)

output_image = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
wb = Workbook()
ws = wb.active
ws.title = "Ergebnisse"

# Überschriften setzen
ws["A1"] = "Flächennummer"
ws["B1"] = "Größe"
ws["C1"] = "Umfang"

# Flächen und Längen berechnen und in Farbe markieren
for i, contour in enumerate(contours):
    area = cv2.contourArea(contour)
    perimeter = cv2.arcLength(contour, True)
    color = generate_random_color()
    cv2.drawContours(output_image, [contour], 0, color, -1)  # -1 füllt die Kontur
    
    # Flächennummer in die Mitte schreiben
    M = cv2.moments(contour)
    if M["m00"] != 0:
        cX = int(M["m10"] / M["m00"])
        cY = int(M["m01"] / M["m00"])
        cv2.putText(output_image, str(i+1), (cX, cY), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 0), 2)
    
    print(f'Fläche Nummer {i+1} - Farbe: {color} - Fläche: {area}, Umfang: {perimeter}')
    
    # Ergebnisse in die Excel-Datei eintragen
    ws.append([i+1, area, perimeter])

# Datei speichern
wb.save("ergebnisse.xlsx")

cv2.imshow('Gefärbte Räume', output_image)
cv2.imwrite("marked.jpg",output_image)
cv2.waitKey(0)
cv2.destroyAllWindows()
